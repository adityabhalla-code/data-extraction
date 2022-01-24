# import libraries
from datetime import datetime
import pandas as pd
import re
from tqdm import tqdm
import warnings
warnings.simplefilter("ignore", UserWarning)
import camelot
import PyPDF2
import PySimpleGUI as sg
from openpyxl import load_workbook


# to display pandas dataframe on max col width
pd.set_option('display.max_colwidth', None)
date_format = "%m/%d/%Y"

# function to divide list into small chunks of list based of n values, here n = 3
def divide_chunks(l, n):
    for i in range(0, len(l), n):
        yield l[i:i + n]

# progress bar visual
def progress_bar(key, iterable, *args, title='', **kwargs):
    """
    Takes your iterable and adds a progress meter onto it
    :param key: Progress Meter key
    :param iterable: your iterable
    :param args: To be shown in one line progress meter
    :param title: Title shown in meter window
    :param kwargs: Other arguments to pass to one_line_progress_meter
    :return:
    """
    sg.one_line_progress_meter(title, 0, len(iterable), key, *args, **kwargs)
    meter = sg.QuickMeter.active_meters[key]
    meter.window.DisableClose = False
    for i, val in enumerate(iterable):
        yield val
        if not sg.one_line_progress_meter(title, i+1, len(iterable), key, *args, **kwargs):
            break

# function to extract the data from pdf
def extract_pdf(file_name):
    result = {}
    # Reading the pdf file by PyPDF2
    reader = PyPDF2.PdfFileReader(file_name)
    # to get number of pages in pdf
    num_pages = reader.getNumPages()
    # getting text from first page of pdf using pypdf2
    firstpage = reader.getPage(0).extractText()
    # TODO: To validate the  format with keyword = AIG LOSS RUN / CNA
    if re.search("AIG\s*Loss\s*Run", firstpage):
        sg.popup_quick_message("valid File format")
        # iterating through all pages
        for page in tqdm(range(num_pages)[5:25], desc="Processing pdf"):
            sg.popup_quick_message('Kindly Wait Processing file.....', background_color='black', text_color='white')
            #sg.one_line_progress_meter("Processing file....",page,float(20),no_button=False)
            #progress_bar(page,range(20),title="Processing file....")
            claim_nos = []
            desc_res_list = []
            text = reader.getPage(page).extractText()
            search = False
            search = re.search("Policy\s*:|Policy\s*Number\s*:", text)
            if search:
                if re.search("Claim\s*#", text):
                    # print(re.search("Claim\s*#",text))
                    # print(page)
                    policy = False
                    search_text = re.split("Policy\s*:\s*[A-Z]+|Policy\s*Number\s*:", text)[-1]
                    policy_pattern = False
                    policy_pattern = re.search("\d{10}-\d{3}-\d{3,4}|[A-Z]+\s*\d{10}\s*[A-Z]+", search_text)
                    # print(policy_pattern)
                    if policy_pattern:
                        # print(policy_pattern)
                        # print(page)
                        policy_pattern = re.search("\d{10}-\d{3}-\d{3,4}|\w+\s*\d{10}", search_text)
                        policy = re.split(policy_pattern.group(0), search_text)[0]
                        policy = policy + " " + policy_pattern.group(0)
                        policy = policy.strip()
                        dates = re.search("\d{2}/\d{2}/\d{4}\s*-\s*\d{2}/\d{2}/\d{4}", search_text).group(0)
                        effective_date = dates.split("-")[0].strip()
                        expiration_date = dates.split("-")[-1].strip()
                        a = datetime.strptime(effective_date, date_format)
                        b = datetime.strptime(expiration_date, date_format)
                        delta = b - a
                        policy_period = delta.days
                        claim_nos = re.findall("\d{3}-\d{6}-\d{3}", search_text)
                        tables = camelot.read_pdf(file_name, pages=str(page + 1), flavor='stream', edge_tol=500, row_tol=1,
                                                  col_tol=1, split_text=True)
                        df = tables[0].df
                        # print(df)
                        row = df[df[0].str.contains("Claimant\s*Name")].index.values[0]
                        # print(row)

                        df = df.drop(df.index[[i for i in range(row)]])
                        df.reset_index(drop=True, inplace=True)
                        reserves = ""
                        recoveries = ""
                        incurred = ""

                        if df[df[0].str.contains("Level\s*3\s*:|Policy\s*:")].shape[0] > 0:
                            end_row = df[df[0].str.contains("Level\s*\d\s*:|Policy\s*:")].index.values[0]
                            df_1 = df.iloc[:end_row, :]
                            # df = df.iloc[:end_row+1:]
                            try:
                                reserves_row = df.iloc[2][df.iloc[2].str.contains("serves")].index.values[0]
                                reserves = df[reserves_row].iloc[-1]
                                if not reserves:
                                    if not df[reserves_row + 1].iloc[0]:
                                        reserves = df[reserves_row + 1].iloc[-1]
                                try:
                                    reserves = reserves.split("\n")[-1]
                                except:
                                    pass
                            except:
                                pass

                            try:
                                recoveries_row = df.iloc[2][df.iloc[2].str.contains("coveries")].index.values[0]
                                recoveries = df[recoveries_row].iloc[-1]
                                if not recoveries:
                                    if not df[recoveries_row + 1].iloc[0]:
                                        recoveries = df[recoveries_row + 1].iloc[-1]
                                try:
                                    recoveries = recoveries.split("\n")[-1]
                                except:
                                    pass
                            except:
                                pass

                            try:
                                incurred_row = df.iloc[2][df.iloc[2].str.contains("curred")].index.values[0]
                                incurred = df[incurred_row].iloc[-1]
                                if not incurred:
                                    if not df[incurred_row + 1].iloc[0]:
                                        incurred = df[incurred_row + 1].iloc[-1]
                                try:
                                    incurred = incurred.split("\n")[-1]
                                except:
                                    pass
                            except:
                                pass
                        else:
                            df_1 = df

                        for i in range(3):  # .index.values[0]
                            v = df_1.iloc[i][df_1.iloc[i].str.contains("Accident|Loss\s*Desc")]
                            if v.to_list():
                                col = v.index.values[0]
                                break

                        desc_list = [desc.split("\n")[-1] for desc in list(df_1[col])]
                        desc_res_list = list(divide_chunks(desc_list, 3))[1:]
                        desc_res_list = [" ".join(desc) for desc in desc_res_list]

                        result[page] = {"policy no": policy, "effective_date": effective_date,
                                        "expiration_date": expiration_date,
                                        "policy_period": policy_period, "claim_nos": claim_nos,
                                        "desc_res_list": desc_res_list,

                                        "total_reserves": reserves, "total_recoveries": recoveries,
                                        "total_incurred": incurred,
                                        # "df":df
                                        }

        #print(result)
        res_df = pd.DataFrame(result).T
        recoveries_indexes = res_df.index[res_df['total_recoveries'] == ""].tolist()
        incurred_indexes = res_df.index[res_df['total_incurred'] == ""].tolist()
        reserves_indexes = res_df.index[res_df['total_reserves'] == ""].tolist()

        if recoveries_indexes == incurred_indexes == reserves_indexes:
            for i in recoveries_indexes:
                if len(res_df) > (i + 1):
                    res_df['total_recoveries'][i] = res_df['total_recoveries'][i + 1]
                    res_df['total_incurred'][i] = res_df['total_incurred'][i + 1]
                    res_df['total_reserves'][i] = res_df['total_reserves'][i + 1]

        for i in res_df['claim_nos'].index.to_list():
            if len(res_df['desc_res_list'][i]) != len(res_df['claim_nos'][i]):
                if re.search('Claim\s*Count', res_df['desc_res_list'][i][-1]):
                    res_df['desc_res_list'][i] = res_df['desc_res_list'][i][:-1]
                if res_df['desc_res_list'][i]:
                    if res_df['desc_res_list'][i][-1].strip().isdecimal():
                        res_df['desc_res_list'][i] = res_df['desc_res_list'][i][:-1]
                if len(res_df['desc_res_list'][i]) != len(res_df['claim_nos'][i]):
                    if re.search('No\s*Claims', res_df['desc_res_list'][i][-1]):
                        res_df['claim_nos'][i] = ["NO CLAIMS" for i in range(len(res_df['desc_res_list'][i]))]
                    else:
                        des_len = len(res_df['desc_res_list'][i])
                        claim_len = len(res_df['claim_nos'][i])
                        if des_len > claim_len:
                            res_df['claim_nos'][i] = res_df['claim_nos'][i] + ["no decs" for i in
                                                                               range(des_len - claim_len)]
                        else:
                            res_df['desc_res_list'][i] = res_df['desc_res_list'][i] + ["no claims" for i in
                                                                                       range(claim_len - des_len)]

        exploded_df = res_df.set_index(res_df.index).apply(pd.Series.explode)
        exploded_df.reset_index(inplace=True, drop=True)
        exploded_df.drop(['total_recoveries', 'total_incurred', 'total_reserves'], axis=1, inplace=True)
        res_df.drop(['desc_res_list'], axis=1, inplace=True)

        return res_df,exploded_df

def write_excel(file_name):
    excel_name = file_name.split(".")[0] + '.xlsx'
    res_df.to_excel(excel_name, index=False)
    exploded_name = file_name.split(".")[0] + '-exploded.xlsx'
    exploded_df.to_excel(exploded_name, index=False)

#### For Excel extraction

# Read the input

def read_excel(file_name,Sheets_flat_list):
    wb = load_workbook(file_name, read_only=True) # open an Excel file and return a workbook
    for i in Sheets_flat_list:
        if i in wb.sheetnames:
            print('sheet1 exists', i)
            Df1 = pd.read_excel(file_name,  sheet_name = i, header=None)
    return Df1


def filter_out(df,entities):
    """
    To filter out extracted table data that match with entities
    :param df: Extracted Table DataFrame
    :param entities: A list  of entities from entity dictionary excel sheet
    :return: Unique indexes of entities matched with excel data

    """
    print(entities)
    found = {}
    df= df.applymap(lambda x : x.lower().strip() if isinstance(x, str) else x)
    #print(df)
    indexes = []
    # Compare keys from extracted data with entities, if keys match with entities then collect information of row, entity, column no
    for i in entities:
        #print(i)
        entity = i
        i=i.lower()
        idf = list(df[df.isin([i.strip()]).any(1)].index)
        column_bool_list = list(df.isin([i.strip()]).any())
        for j in range(len(column_bool_list)):
            if column_bool_list[j] == True:
                column_no = j
        if len(idf)>0:

            indexes.append((idf[0],entity,column_no))

    return indexes


def find_sub_set_df(df, index):
    dictionary = {}
    # If entity is at start of Dataframe then take values from 0 to end of dataframe
    for i in index:
        sub_set = df[i[-1]][i[0]:len(df)]
        dictionary[i[1]] = sub_set.to_list()[1:]
    df = pd.DataFrame.from_dict(dictionary, orient = 'index')
    return dictionary,df.T


def Policyterm_to_cols(Policyterm,carrier_name, df):
    """
    Split date and add as columns to df
    TODO: Test on CNA
    """
    temp = []
    effect_date = []
    expiration_date = []
    if carrier_name == "AIG":
        for i in Policyterm:
            temp = re.findall('\d{2}/\d{2}/\d{4}', i)
            #print("temp",temp)
            effect_date.append(temp[0])
            expiration_date.append(temp[1])
        df["effective date"] = effect_date
        df["expiration date"] = expiration_date
        df.drop('Policy Term', axis=1, inplace=True)
    if carrier_name == "CNA":
        for i in Policyterm:
            temp = re.findall('\d{2}/\d{2}/\d{2}', i)
            #print("temp",temp)
            effect_date.append(temp[0])
            expiration_date.append(temp[1])
        df["effective date"] = effect_date
        df["expiration date"] = expiration_date
        df.drop("Eff - Exp Dates",axis=1,inplace=True)
    return effect_date,expiration_date ,df
