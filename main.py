
from Utility import *
from openpyxl import load_workbook

# importing dictionary of features

# lossrun_sheet_names > names of the tabs to extract data from carieer wise

# sheet_1 > all column names as per the excel format provided
# sheet 1 in loss run dictionary
try:
    filename = 'output_gui.csv'
    dict_name = "Lossrun Dict (1).xlsx"
    df = pd.read_csv(filename)
    # TODO: change variable names for below sheets
    Sheets = pd.read_excel(r"Lossrun Dict (1).xlsx", sheet_name="Lossrun_sheet_names")
    entities = pd.read_excel(r"Lossrun Dict (1).xlsx", sheet_name="Sheet1")
    dummy_data = df.values.tolist()
    #headings = list(df.columns)
except Exception as e:
    print("Error reading data")
    print(traceback.format_exc())
    pass


# to display pandas dataframe on max col width
pd.set_option('display.max_colwidth', None)
date_format = "%m/%d/%Y"

# predefined variable names
headings = ["policy no","effective_date" ,"expiration_date","policy_period","claim_nos","desc_res_list","total_reserves","total_recoveries","total_incurred"]

# reaaranging the columns as per the final format
excel_headings = ["Policy no", "effective date","expiration date","Claim no","OneClaim no","Loss Description","Total Recoveries","Total Incurred","Total Reserves"]
# Column name dictionary for AIG  output
col_rename_AIG = {'Policy-Asco-Mod':'Policy no', 'Claim #': 'Claim no', "OneClaim #":"OneClaim no",
                     "Total \nRecoveries":"Total Recoveries","Total \nIncurred":"Total Incurred",
                      "Total \nReserves":"Total Reserves","Accident / Loss Description":"Loss Description"}
# Column name dictionary for CNA output
col_rename_CNA = {"Eff - Exp Dates":"Policy Period","Paid Expenses":"Total Paid Expenses",
                     "Reserves":"Total Reserves","Recovery":"Total Recovery"}


# GUI Initial settings
sg.ChangeLookAndFeel('Reddit')
sg.popup_quick_message('Loading your GUI.... one moment..', background_color='black', text_color='white')
font = ("Arial", 10)
attribute_font = ("Arial", 15)

## Layput
opening_page_layout = [
    # [sg.Text('Entities Extractor', font=font,auto_size_text=True)],
    [sg.Text(
        "1. Select the file\n2. Click on extract.\n3. New Window for extracted data opens.\n4. Select the output path\n5. Click save to save the file",
        font=font)],
    [sg.In('file path', key='-IN-PATH-', enable_events=True), sg.FileBrowse()],
    [sg.Button('Extract'), sg.Button('Cancel')]
]

data_view_layout = [
    [sg.Text("Extracted Data", font=attribute_font)],
    [sg.Table(values=dummy_data,
              headings=headings,
              key='table',
              display_row_numbers=True,
              auto_size_columns=False,
              num_rows=min(25, len(dummy_data)))],
    [sg.In('Path to save Data', size=(40, 10), key='-OUT-PATH-'), sg.FolderBrowse()],
    [sg.Button("Save"), sg.Button('Back'), sg.Button("Exploded View"),sg.Stretch()],
]

exploded_view_layout = [
    [sg.Text("Policy wise data", font=attribute_font)],
    [sg.Table(values=dummy_data,
              headings=headings,
              key='table-2',
              display_row_numbers=True,
              auto_size_columns=False,
              num_rows=min(25, len(dummy_data)))],
    [sg.In('Path to save Data', size=(40, 10), key='-OUT-PATH-'), sg.FolderBrowse()],
    [sg.Button("Save",enable_events=True,key="save"), sg.Button('Back',enable_events=True,key="back")],
]

# entities_layout = [[sg.Table(values=data,
#                   headings=headings,
#                   key='table',
#                   display_row_numbers=True,
#                   auto_size_columns=False,
#                   num_rows=min(25, len(data)))]]


# ----------- Create actual layout using Columns and a row of Buttons
# layout = [[sg.Column(layout1, key='-COL1-'), sg.Column(layout2, visible=False, key='-COL2-'), sg.Column(layout3, visible=False, key='-COL3-')],
#           [sg.Button('Cycle Layout'), sg.Button('1'), sg.Button('2'), sg.Button('3'), sg.Button('Exit')]]

# TOOD:  pandas profile layput
full_layout = [[sg.Column(opening_page_layout, key='-COL1-'),
                sg.Column(data_view_layout, visible=False, key='-COL2-'),
                sg.Column(exploded_view_layout,visible=False,key='-COL3-')]]  # ,
#                sg.VSeparator(),
#                sg.Column(entities_layout)]]


win = sg.Window('Entities Extractor',
                default_element_size=(40, 6),
                text_justification='left',
                auto_size_text=False).Layout(full_layout)

# print(win['table'])
while True:
    event, values = win.Read()
    if event == sg.WIN_CLOSED or event == 'Cancel':  # or event == 'Ok':
        break

    if event == "Back":
        win['-COL3-'].update(visible=False)
        win['-COL2-'].update(visible=False)
        win['-COL1-'].update(visible=True)


    if event == "Save":
        output_file_path = values["-OUT-PATH-"]
        if output_file_path == "Path to save Data":
            sg.popup_quick_message('No output path selected.....', background_color='black', text_color='white')
            pass
        else:
            # path = output_file_path + "/"
            # excel_name = file_name.split(".")[0] + '.xlsx'
            # res_df.to_excel(excel_name, index=False)
            # exploded_name = file_name.split(".")[0] + '-exploded.xlsx'
            # exploded_df.to_excel(exploded_name, index=False)
            write_excel(file_name)

    if event == "Extract":
        if values['-IN-PATH-'] == "file path":
            sg.popup_quick_message("Please select the input file", background_color="black", text_color="white")
            pass
        else:
            # win['-COL2-'].update(visible=True)
            file_name = values['-IN-PATH-']  # .split("/")[-1]
            print(file_name)
            file_extension = file_name[-4:]
            if file_extension == ".pdf":
                res_df, exploded_df = extract_pdf(file_name)
                data = res_df.values.tolist()
                headings = res_df.columns
                #                 table = sg.Table(values=data,
                #                       headings=headings,
                #                       key='-TABLE-',
                #                       display_row_numbers=True,
                #                       auto_size_columns=False,
                #                       num_rows=min(25, len(data)))
                #                 win.Element("-DATA-").Update(res_df)

                win['table'].update(values=data)
                #                win['table'].update(headings=headings)
                win['-COL2-'].update(visible=True)
                win['-COL1-'].update(visible=False)

            elif file_extension == "xlsx" or "XLSX":
                excel_file_name = file_name.split("/")[-1].split(".")[0]
                print(excel_file_name)
                #TODO: Integrate the code here and variable name change
                carrier_name = excel_file_name.split(" ")[0]
                print(carrier_name)
                #input_file_name = (excel_file_name.split(".xlsx")[0]).split("\\")[-1]
                #print(input_file_name)
                # Sheets : identifier of required sheet name in input file
                # entities: identifier of the required output features on the identified Sheets
                Sheets = Sheets[carrier_name].to_frame()
                Sheets_list = Sheets.values.tolist()
                #print(Sheets_list)
                Sheets_flat_list = [str(item).strip() for sublist in Sheets_list for item in sublist]
                #print(Sheets_flat_list)
                entities = entities[carrier_name].to_frame()
                #print(entities)
                entity_list = entities.values.tolist()
                #print(entity_list)
                entity_flat_list = [item.strip() for sublist in entity_list for item in sublist if isinstance(item, str)]
                #print(entity_flat_list)
                #file = file_name+file_extension
                wb = load_workbook(file_name, read_only=True)  # open an Excel file and return a workbook
                for i in Sheets_flat_list:
                    if i in wb.sheetnames:
                        print('sheet1 exists', i)
                        Df1 = pd.read_excel(file_name, sheet_name=i, header=None)
                # data > lst
                lst = filter_out(Df1,entity_flat_list)
                print(lst)
                a, Df2 = find_sub_set_df(Df1, lst)
                #Df2.head(5)
                if carrier_name == "AIG":
                    _, _, Df2 = Policyterm_to_cols(Df2['Policy Term'],carrier_name, Df2)
                    Df2.rename(columns=col_rename_AIG, inplace=True)
                if carrier_name == "CNA":
                    _,_,Df2 = Policyterm_to_cols(Df2['Eff - Exp Dates'],carrier_name,Df2)
                    Df2.rename(columns=col_rename_CNA, inplace=True)
                # reaaranging the columns as per the final format
                # TODO: files where all columns are not found
                #Df2 = Df2[excel_headings]
                data = Df2.values.tolist()
                headings = Df2.columns
                #                 table = sg.Table(values=data,
                #                       headings=headings,
                #                       key='-TABLE-',
                #                       display_row_numbers=True,
                #                       auto_size_columns=False,
                #                       num_rows=min(25, len(data)))
                #                 win.Element("-DATA-").Update(res_df)

                win['table'].update(values=data)
                #                win['table'].update(headings=headings)
                win['-COL2-'].update(visible=True)
                win['-COL1-'].update(visible=False)

    if event =="Exploded View":
        data = exploded_df.values.tolist()
        headings = exploded_df.columns
        win['table-2'].update(values=data)
        win['-COL3-'].update(visible=True)
        #win['-COL2-'].update(visible=False)

    # if win['back'] is not None:
    #     win['-COL3-'].update(visible=False)
    #     win['-COL2-'].update(visible=True)

#     if event =="Save":


win.close()
